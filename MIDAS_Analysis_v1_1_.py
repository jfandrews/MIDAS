
    # -*- coding: utf-8 -*-
"""
Created on Wed Jan 31 13:20:49 2018

@author: Joel F. Andrews
"""
import logging
import sys
import codecs
from PyQt5 import QtWidgets
import numpy as np
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.backends.backend_qt5agg import NavigationToolbar2QT as NavigationToolbar
import matplotlib.pyplot as plt
import openpyxl as xl
from scipy import stats
import os.path
import math
import easygui as eg

logging.basicConfig(filename='MIDAS.log',level=logging.DEBUG)
logging.debug('program started')

class Window(QtWidgets.QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.Open()
        #logging.debug('Open() called')
        
        self.figure = plt.figure()
        self.ax = self.figure.add_subplot(111)
        self.canvas = FigureCanvas(self.figure)
        self.toolbar = NavigationToolbar(self.canvas, self)
        
        
            #groupboxes
        self.viewbox = QtWidgets.QGroupBox('Data View')
        self.roibox = QtWidgets.QGroupBox('Included ROIs')
        self.databox = QtWidgets.QGroupBox('Data Source')
        self.ctrlbox = QtWidgets.QGroupBox('Control ROIs')
        self.plotbox = QtWidgets.QGroupBox('Plotting Options')  
        self.plot_optionsbox = QtWidgets.QGroupBox('Plotting Options')
        self.summarybox = QtWidgets.QGroupBox('Data Summary')
        self.data_optionsbox = QtWidgets.QGroupBox('Data Source')
        
            #comboboxes 
        self.cb = QtWidgets.QComboBox()
        self.cb.addItems(['Data','Controls'])
        self.cb.currentIndexChanged.connect(self.selectionchange)
        self.cb2 = QtWidgets.QComboBox()
        self.cb2.addItems(['Plot Lines', 'Plot Mean'])
        self.cb2.currentIndexChanged.connect(self.selectionchange2)
        
            #buttons
        self.buttons = [QtWidgets.QRadioButton('Stimulation ROI Intensity'), QtWidgets.QRadioButton('Reference ROI Intensity'),QtWidgets.QRadioButton('Reference Division'),QtWidgets.QRadioButton('Reference Subtraction'),QtWidgets.QRadioButton('Normalized to Max'),QtWidgets.QRadioButton('Normalized to First Frame')]
        self.schemeNames = ['RawStim', 'RawRef', 'Normalized to Reference', 'Normalized to Max', 'Normalized to First Frame']  
        self.buttongroup = QtWidgets.QButtonGroup()
        self.ROIbuttongroup = QtWidgets.QButtonGroup()
        self.plotbuttongroup = QtWidgets.QButtonGroup()
        self.controlbuttongroup = QtWidgets.QButtonGroup()
        self.xlbutton = QtWidgets.QPushButton('Export to Excel')
        self.xlbutton.clicked.connect(self.xl_button_clicked)
        self.PlotButtons = [QtWidgets.QRadioButton('Plot Lines'), QtWidgets.QRadioButton('Plot Binned')]
        self.ControlButtons = [QtWidgets.QRadioButton('Data'), QtWidgets.QRadioButton('Controls')]
        
            #layouts
        self.layout = QtWidgets.QVBoxLayout() #main layout
        self.data_options = QtWidgets.QHBoxLayout()
        self.plot_options = QtWidgets.QHBoxLayout()
        self.norm_options = QtWidgets.QHBoxLayout()
        self.options = QtWidgets.QHBoxLayout()
        summary = QtWidgets.QVBoxLayout()
        self.layout = QtWidgets.QVBoxLayout()
        self.layout2 = QtWidgets.QHBoxLayout()
        #layoutControl = QtWidgets.QHBoxLayout()
        self.layoutSplit = QtWidgets.QHBoxLayout()
        layout3 = QtWidgets.QHBoxLayout()
        #layout4 = QtWidgets.QHBoxLayout()
        layoutView = QtWidgets.QVBoxLayout()
            #labels
        self.namebox = QtWidgets.QLabel()
        self.ROItext = QtWidgets.QLabel()
        self.textbox = QtWidgets.QLabel() 
        self.textbox2 = QtWidgets.QLabel()
        self.textbox3 = QtWidgets.QLabel()
            #all layout operations after this point
            
        self.layout.addWidget(self.toolbar)
        self.layout.addWidget(self.canvas)
        
        self.data_options.addWidget(self.cb)
        self.plot_options.addWidget(self.cb2)
        self.data_optionsbox.setLayout(self.data_options)
        self.plot_optionsbox.setLayout(self.plot_options)
        self.options.addWidget(self.data_optionsbox)
        self.options.addWidget(self.plot_optionsbox)
        self.layout.addLayout(self.options)
        for i in range(len(self.buttons)):
            layoutView.addWidget(self.buttons[i])
            self.buttongroup.addButton(self.buttons[i],i+1)
            self.buttons[i].clicked.connect(self.radio_button_clicked)
        self.buttons[3].setChecked(True) 
        self.scheme = self.buttongroup.checkedId()
        self.viewbox.setLayout(layoutView)
        layout3.addWidget(self.viewbox)
        summary.addWidget(self.namebox)
        summary.addWidget(self.textbox)
        summary.addWidget(self.textbox2)
        summary.addWidget(self.textbox3)
        self.summarybox.setLayout(summary)
        layout3.addWidget(self.summarybox)
        self.layout.addLayout(layout3)
        RoiBoxes = ['']*len(self.dataFiles)
        CtrlBoxes = ['']*len(self.dataFiles)
        RoiLayouts = ['']*len(self.dataFiles)
        CtrlLayouts = ['']*len(self.dataFiles)
        self.Chk = []
        for dF in range(len(self.dataFiles)):
            x = str(dF+1)
            RoiBoxes[dF] = 'Roi Box' + x
            RoiBoxes[dF] = QtWidgets.QGroupBox('DataSet_' + x)
            RoiLayouts[dF] = 'Roi Layout' + x
            RoiLayouts[dF] = QtWidgets.QHBoxLayout()
            CtrlBoxes[dF] = 'Ctrl Box' + x
            CtrlBoxes[dF] = QtWidgets.QGroupBox('Controls_' + x)
            CtrlLayouts[dF] = 'Ctrl Layout' + x
            CtrlLayouts[dF] = QtWidgets.QHBoxLayout()
            self.Chk.append([])
            
            for z in range(self.dmgRois[dF]):
                
                y = str(z+1)
                self.Chk[dF].append('self.checkbox' + y)
                #self.Chk[dF][z] = 'self.checkbox' + y                    
                self.Chk[dF][z] = QtWidgets.QCheckBox(y)
                if (roiInfo[dF][z][0] == 0.0):
                    CtrlLayouts[dF].addWidget(self.Chk[dF][z])
                    self.Chk[dF][z].setChecked(False)
                
                else:
                    RoiLayouts[dF].addWidget(self.Chk[dF][z])
                    self.Chk[dF][z].setChecked(True)            
                self.Chk[dF][z].clicked.connect(self.btnstate)         
            RoiBoxes[dF].setLayout(RoiLayouts[dF])
            CtrlBoxes[dF].setLayout(CtrlLayouts[dF])
        for dF in range(len(self.dataFiles)):
            self.layout.addWidget(RoiBoxes[dF])
            self.layout.addWidget(CtrlBoxes[dF])
        self.layout.addWidget(self.xlbutton)
        self.cb.setCurrentIndex(0)
        self.controlType = self.cb.currentIndex()
        self.cb2.setCurrentIndex(1)
        self.plotType = self.cb2.currentIndex()
        self.setLayout(self.layout)  
        
        self.plot()
        logging.debug('__init__() Completed')
        
    def AddData(self):
        dF = QtWidgets.QFileDialog.getOpenFileNames(self, 'Select Data File')[0]
        logging.debug('Data File Selected')
        print(dF)
        for d in range(len(dF)):
            baseName = os.path.basename(dF[d])
            nameLength = len(baseName)
            path = os.path.normpath(dF[d])[:-nameLength]
            self.filenames.append(baseName[:-4])
        
            self.paths.append(path)
            self.dataFiles.append(dF[d])
            setname = dF[d][:-12]+'_Settings.txt'
            print(setname)
            if os.path.exists(setname):
                print('Settings found for ' + baseName)
                sF = setname
            else:
                for s in range(len(dF)):
                    sF = QtWidgets.QFileDialog.getOpenFileName(self, 'Select Settings File for ' + self.filenames[s])[0]
    
            self.settingsFiles.append(sF)
        
        add = QtWidgets.QMessageBox.question(None,'Add DataSets','Add Another Dataset?',QtWidgets.QMessageBox.Yes,QtWidgets.QMessageBox.No)
        if add == QtWidgets.QMessageBox.Yes:
            
            self.AddData()
        else:
            return len(self.dataFiles)
        
        logging.debug('AddData() Completed')
        
    def Open(self):
        self.dataFiles = []
        self.settingsFiles = []
        self.dmgRois = []
        self.interval = []
        
        self.dataStructure = {'Frames':0,'rawStim':1,'refStim':2,'refDiv':3,'refSub':4,'normMax':5,'normFirst':6,'Time':7}
        global controlNumbers
        controlNumbers = []
        
        global data 
        data = []
        global controlData
        controlData = []
        
        global roiInfo
        roiInfo = []  
        self.frames = []
        self.filenames = []
        self.paths = []
        self.AddData()
        self.experimentGroup = QtWidgets.QInputDialog.getText(self,'User Input', 'Name of Experimental Group?')
        logging.debug('User Input Group Name')
        
        filecount=0
        for dF in range(len(self.dataFiles)):
            filecount=filecount+1
            controlNumber = 0
            with codecs.open(self.dataFiles[dF], encoding = 'utf-16-le') as dfile:
                colnames = dfile.readlines()[0].strip('\n').split('\t')
            with codecs.open(self.dataFiles[dF], encoding = 'utf-16-le') as dfile:
                d = np.loadtxt(dfile, skiprows = 1, dtype = {'names': colnames,'formats':['float' for c in colnames]}, delimiter = '\t')
            g = np.sort(d,order = ['ROI ID','Time'])
            maxrow = len(g)
            maxROI = g[maxrow-1][0]
            self.dmgRois.append(np.int(maxROI/2))
            print(self.dmgRois[dF])
            
            
          
            
            settings = codecs.open(self.settingsFiles[dF], 'r', encoding = 'utf-16-le')
            settingslines = settings.readlines()
            last = len(settingslines)
            self.interval.append(float(settingslines[1].split('\t')[3]))
            maxTimeOffset = float(settingslines[last-1].split('\t')[2])
            
            
                
            
            
            
            print(self.interval[dF],maxTimeOffset)
            print(float(settingslines[3].split('\t')[1]))
            
            
            
            
             
            roiInfo.append(np.zeros((self.dmgRois[dF],6)))
            for q in range(self.dmgRois[dF]):
                
                roiInfo[dF].itemset((q,0),float(settingslines[3+q].split('\t')[1])) 
                print(roiInfo[dF][q][0])      #adds the power level of each ROI to the self.roiInfo
                if (roiInfo[dF][q][0] ==0.0):
                    controlNumber = controlNumber + 1   #counts the number of controls
                    print('control found!')
                roiInfo[dF].itemset((q,1),float(settingslines[3+q].split('\t')[2]))  #adds the offset of each ROI to RoiInfo
            
            controlNumbers.append(controlNumber)
            self.frames.append(np.int(maxrow/maxROI))
            offset = np.int(maxrow/2)
            data.append(np.zeros((self.frames[dF],8,self.dmgRois[dF])))
            if filecount==1:
                sel = eg.choicebox('Select Data Source','Data Selection', colnames[2:])
            l = 0
            for m in range(self.dmgRois[dF]):
                for n in range(self.frames[dF]):
                    data[dF].itemset((n,self.dataStructure['Frames'],m),g[l][1])                          #time
                    data[dF].itemset((n,self.dataStructure['rawStim'],m),g[l][sel])                          #stimROI
                    data[dF].itemset((n,self.dataStructure['refStim'],m),g[l+offset][sel])                   #refROI
                    data[dF].itemset((n,self.dataStructure['refSub'],m),g[l][sel]-g[l+offset][sel]) 
                    data[dF].itemset((n,self.dataStructure['refDiv'],m),g[l][sel]/g[l+offset][sel])
                    
                    l = l+1
        
            
            for o in range(self.dmgRois[dF]):
                normMax = data[dF][:,self.dataStructure['rawStim'],o].max()
                maxindex = data[dF][:,self.dataStructure['rawStim'],o].argmax(axis=0)
                normMin = data[dF][:,self.dataStructure['rawStim'],o].min()
                normFirst = data[dF][0,self.dataStructure['rawStim'],o]
                roiInfo[dF].itemset((o,4),normMax/normFirst)    
                
                for p in range(self.frames[dF]):
                    data[dF].itemset((p,self.dataStructure['normMax'],o),(data[dF][p][1][o]-normMin)/(normMax-normMin))
                    data[dF].itemset((p,self.dataStructure['normFirst'],o),data[dF][p][1][o]/normFirst)
                    
            for q in range(self.dmgRois[dF]):
                data[dF].itemset((0,0,q),0)                                  #set first frame time to 0, all other times will be offset from it
                for r in range(1,self.frames[dF]):
                    data[dF].itemset((r,self.dataStructure['Time'],q),data[dF][r][0][q]+maxTimeOffset-roiInfo[dF][q][1])    #each ROI has the max time offset applied to it and each ROI's individual 
                maxindex = data[dF][:,1,q].argmax(axis=0)
                
                for s in range(self.frames[dF]):
                    if(data[dF][s,self.dataStructure['normMax'],q]<0.95):
                        continue
                    else:
                        roiInfo[dF].itemset((q,2),data[dF][s,self.dataStructure['Time'],q]/1000)
                        break
                for t in range(maxindex+1 ,self.frames[dF]):
                    if(data[dF][t,self.dataStructure['normMax'],q]>0.55):
                        continue
                    else:
                        roiInfo[dF].itemset((q,3),data[dF][t,self.dataStructure['Time'],q]/1000)
                        break        
                    
           
        logging.debug('Open() Completed')
        
    def xl_button_clicked(self):                                    #creates excel spreadsheet showing binned values from included ROIs 
       
        wb = xl.Workbook()
        wb.create_sheet(title = 'Summary')
        summary = wb['Summary']
        summary.cell(row =1, column = 1).value = 'Included Files:'
        for dF in range(len(self.dataFiles)):
            summary.cell(row =1, column = 2+dF).value = self.filenames[dF][:-5] + '.nd2'
        summary.cell(row =2, column = 1).value = 'Mean Recruitment Peak Time'
        summary.cell(row =3, column = 1).value = self.peakTime
        summary.cell(row =2, column = 2).value = 'Recruitment Peak SEM'
        summary.cell(row =3, column = 2).value = self.peakError
        summary.cell(row =2, column = 3).value = 'Mean Recruitment Half-Life'
        summary.cell(row =3, column = 3).value = self.halfTime
        summary.cell(row =2, column = 4).value = 'Half-Life SEM'
        summary.cell(row =3, column = 4).value = self.halfError
        summary.cell(row = 4, column = 1).value = 'Filename'
        summary.cell(row = 4, column = 2).value = 'ROI'
        summary.cell(row = 4, column = 3).value = 'Power Level'
        summary.cell(row = 4, column = 4).value = 'Timing Offset'
        summary.cell(row = 4, column = 5).value = 'Peak Recruitment Time'
        summary.cell(row = 4, column = 6).value = 'Recruitment Half-Life'
        summary.cell(row = 4, column = 7).value = 'Relative Peak Intensity'
        summary.cell(row = 4, column = 8).value = 'Included in Analysis?'
       
        l=5
        for dF in range(len(self.dataFiles)):
            for r in range(self.dmgRois[dF]):
                summary.cell(row = r+l ,column = 1).value = self.filenames[dF][:-5] + '.nd2'
                summary.cell(row = r+l ,column = 2).value = r+1
                summary.cell(row = r+l,column = 3).value = roiInfo[dF][r][0]
                summary.cell(row = r+l,column = 4).value = roiInfo[dF][r][1]
                summary.cell(row = r+l,column = 5).value = roiInfo[dF][r][2]
                summary.cell(row = r+l,column = 6).value = roiInfo[dF][r][3]
                summary.cell(row = r+l,column = 7).value = roiInfo[dF][r][4]
                if self.Chk[dF][r].isChecked() == True:
                    summary.cell(row = r+l,column = 8).value = 'Yes'
                else:
                    summary.cell(row = r+l,column = 8).value = 'No'
            l = l+ self.dmgRois[dF]
         
        wb.create_sheet(title = 'Mean_Data')
        
        means = wb['Mean_Data']
        means.cell(row = 1, column = 1).value = 'Time'
        means.cell(row = 1, column = 2).value = 'Stimulation Intensity'
        means.cell(row = 1, column = 3).value = 'Stim SEM'
        means.cell(row = 1, column = 4).value = 'Reference Intensity'
        means.cell(row = 1, column = 5).value = 'Ref SEM'
        means.cell(row = 1, column = 6).value = 'Stim/Ref'
        means.cell(row = 1, column = 7).value = 'Stim/Ref SEM'
        means.cell(row = 1, column = 8).value = 'Stim-Ref'
        means.cell(row = 1, column = 9).value = 'Stim-Ref SEM'
        means.cell(row = 1, column = 10).value = 'Normalized to Maximum'
        means.cell(row = 1, column = 11).value = 'NormMax SEM'
        means.cell(row = 1, column = 12).value = 'Normalized to First Frame'
        means.cell(row = 1, column = 13).value = 'NormFirst SEM'
        
        xlTime = np.arange(0,self.binMeans.size*max(self.interval),max(self.interval))
        xlMeans = []
        meansSEM = []
        for m in range(6):
            xlMeans.append([])
            meansSEM.append([])
        
        xlMeans[0] = stats.binned_statistic(self.binData[0],self.binData[self.dataStructure['rawStim']], bins = max(self.frames)-1)[0]
        xlMeans[1] = stats.binned_statistic(self.binData[0],self.binData[self.dataStructure['refStim']], bins = max(self.frames)-1)[0]
       
        xlMeans[2] = stats.binned_statistic(self.binData[0],self.binData[self.dataStructure['refDiv']], bins = max(self.frames)-1)[0]
        xlMeans[3] = stats.binned_statistic(self.binData[0],self.binData[self.dataStructure['refSub']], bins = max(self.frames)-1)[0]
        xlMeans[4] = stats.binned_statistic(self.binData[0],self.binData[self.dataStructure['normMax']], bins = max(self.frames)-1)[0]
        xlMeans[5] = stats.binned_statistic(self.binData[0],self.binData[self.dataStructure['normFirst']], bins = max(self.frames)-1)[0]
       
        meansSEM[0] = stats.binned_statistic(self.binData[0],self.binData[self.dataStructure['rawStim']], statistic = lambda z: stats.sem(z), bins = max(self.frames)-1)[0]  
        meansSEM[1] = stats.binned_statistic(self.binData[0],self.binData[self.dataStructure['refStim']], statistic = lambda z: stats.sem(z), bins = max(self.frames)-1)[0]
        meansSEM[2] = stats.binned_statistic(self.binData[0],self.binData[self.dataStructure['refDiv']], statistic = lambda z: stats.sem(z), bins = max(self.frames)-1)[0]
        meansSEM[3] = stats.binned_statistic(self.binData[0],self.binData[self.dataStructure['refSub']], statistic = lambda z: stats.sem(z), bins = max(self.frames)-1)[0]
        meansSEM[4] = stats.binned_statistic(self.binData[0],self.binData[self.dataStructure['normMax']], statistic = lambda z: stats.sem(z), bins = max(self.frames)-1)[0]
        meansSEM[5] = stats.binned_statistic(self.binData[0],self.binData[self.dataStructure['normFirst']], statistic = lambda z: stats.sem(z), bins = max(self.frames)-1)[0]
        
        
        
        
        
        
        for t in range(len(xlTime)):
            means.cell(row = t+2, column =1).value = xlTime[t]
            
            means.cell(row = t+2, column =2).value = float(format(xlMeans[0][t], '.2f'))
            means.cell(row = t+2, column =3).value = float(format(meansSEM[0][t], '.2f'))
            means.cell(row = t+2, column =4).value = float(format(xlMeans[1][t], '.2f'))
            means.cell(row = t+2, column =5).value = float(format(meansSEM[1][t], '.2f'))
            means.cell(row = t+2, column =6).value = float(format(xlMeans[2][t], '.2f'))
            means.cell(row = t+2, column =7).value = float(format(meansSEM[2][t], '.2f'))
            means.cell(row = t+2, column =8).value = float(format(xlMeans[3][t], '.2f'))
            means.cell(row = t+2, column =9).value = float(format(meansSEM[3][t], '.2f'))
            means.cell(row = t+2, column =10).value =float(format(xlMeans[4][t], '.2f'))
            means.cell(row = t+2, column =11).value =float(format(meansSEM[4][t], '.2f'))
            means.cell(row = t+2, column =12).value =float(format(xlMeans[5][t], '.2f'))
            means.cell(row = t+2, column =13).value =float(format(meansSEM[5][t], '.2f'))
        l = 2    
        for dF in range(len(self.dataFiles)):
            for r in range(self.dmgRois[dF]):
                
                title = 'File_ ' + str(dF+1) + '_ROI_' + str(r+1)
                wb.create_sheet(title = title)
                sheet = wb[title]
                sheet.cell(row=1,column=1).value = self.filenames[dF][:-5]+'.nd2'
                sheet.cell(row = 2, column = 1).value = 'Time'
                sheet.cell(row = 2, column = 2).value = 'Stimulation Spot Intensity'
                sheet.cell(row = 2, column = 3).value = 'Reference Spot Intensity'
                sheet.cell(row = 2, column = 4).value = 'Reference Corrected by Division'
                sheet.cell(row = 2, column = 5).value = 'Reference Corrected by Subtraction'
                sheet.cell(row = 2, column = 6).value = 'Normalized to Maximum'
                sheet.cell(row = 2, column = 7).value = 'Normalized to First Frame'
                        
                for f in range(self.frames[dF]):
                    sheet.cell(row = f+3, column =1).value = data[dF][f][self.dataStructure['Time']][r]
                    sheet.cell(row = f+3, column =2).value = data[dF][f][self.dataStructure['rawStim']][r]
                    sheet.cell(row = f+3, column =3).value = data[dF][f][self.dataStructure['refStim']][r]
                    sheet.cell(row = f+3, column =4).value = data[dF][f][self.dataStructure['refDiv']][r]
                    sheet.cell(row = f+3, column =5).value = data[dF][f][self.dataStructure['refSub']][r]
                    sheet.cell(row = f+3, column =6).value = data[dF][f][self.dataStructure['normMax']][r]
                    sheet.cell(row = f+3, column =7).value = data[dF][f][self.dataStructure['normFirst']][r]
                l=l+1
                    
                    
        
        
     
     
        wb.remove_sheet(wb.worksheets[0])
        wb.save(self.paths[0] +  self.experimentGroup[0] + '_Analysis.xlsx')
        logging.debug('XL file created')          
       
    def plot(self):
        logging.debug('Plot() begun')
        count = 0
        self.binData = []
        self.binData.append([])
        self.binData.append([])
        self.binData.append([])
        self.binData.append([])
        self.binData.append([])
        self.binData.append([])
        self.binData.append([])
        
        self.binControl = []
        self.binControl.append([])
        self.binControl.append([])
        self.binControl.append([])
        self.binControl.append([])
        self.binControl.append([])
        self.binControl.append([])
        self.binControl.append([])
        
        for dF in range(len(self.dataFiles)):
            for s in range(self.dmgRois[dF]):                     #if ROIs are checked, adds values to selectedData
                    
                if self.Chk[dF][s].isChecked() == True:
                    count = count + 1
                    for f in range(self.frames[dF]):
                        self.binData[0].append(data[dF][f,self.dataStructure['Time'],s])
                        self.binData[1].append(data[dF][f,self.dataStructure['rawStim'],s])  
                        self.binData[2].append(data[dF][f,self.dataStructure['refStim'],s])  
                        self.binData[3].append(data[dF][f,self.dataStructure['refDiv'],s])  
                        self.binData[4].append(data[dF][f,self.dataStructure['refSub'],s])  
                        self.binData[5].append(data[dF][f,self.dataStructure['normMax'],s])  
                        self.binData[6].append(data[dF][f,self.dataStructure['normFirst'],s])
                        
                    
                if roiInfo[dF][s][0] ==0.0:
                    for f in range(self.frames[dF]):
                        self.binControl[0].append(data[dF][f,self.dataStructure['Time'],s])
                        self.binControl[1].append(data[dF][f,self.dataStructure['rawStim'],s])  
                        self.binControl[2].append(data[dF][f,self.dataStructure['refStim'],s])  
                        self.binControl[3].append(data[dF][f,self.dataStructure['refDiv'],s])  
                        self.binControl[4].append(data[dF][f,self.dataStructure['refSub'],s])  
                        self.binControl[5].append(data[dF][f,self.dataStructure['normMax'],s])#append corrected time to self.binSource[0]
                        self.binControl[6].append(data[dF][f,self.dataStructure['normFirst'],s])
                        
                        
        self.binMeans = stats.binned_statistic(self.binData[0],self.binData[self.scheme], bins = max(self.frames)-1)[0]
        self.binSEM =   stats.binned_statistic(self.binData[0],self.binData[self.scheme], statistic = lambda z: stats.sem(z), bins = max(self.frames)-1)[0]   
        if sum(x for x in controlNumbers) != 0:
            self.controlMeans = stats.binned_statistic(self.binControl[0],self.binControl[self.scheme], bins = max(self.frames)-1)[0]
            self.controlSEM =   stats.binned_statistic(self.binControl[0],self.binControl[self.scheme], statistic = lambda z: stats.sem(z), bins = max(self.frames)-1)[0]
        self.peaks()
        if self.controlType ==0:
            
            if self.plotType ==0:
                for dF in range(len(self.dataFiles)):
                    for s in range(self.dmgRois[dF]):
                            if self.Chk[dF][s].isChecked()==True:
                                x = data[dF][:,self.dataStructure['Time'],s]/1000
                                y = data[dF][:,self.scheme,s]
                                self.ax.plot(x,y)
                                self.axes()
                                self.canvas.draw()
                            else:continue
            if self.plotType ==1:
                for dF in range(len(self.dataFiles)):
                    
                    x = np.arange(0,self.binMeans.size*max(self.interval),max(self.interval))
                    y = self.binMeans
                    plt.errorbar(x,y,color = 'b', yerr = self.binSEM)
                    self.axes()
                    self.canvas.draw()
                    
        if self.controlType ==1:
            if self.plotType ==0:
                for dF in range(len(self.dataFiles)):
                    for s in range(self.dmgRois[dF]):
                            if roiInfo[dF][s][0] == 0:
                                x = data[dF][:,self.dataStructure['Time'],s]/1000
                                y = data[dF][:,self.scheme,s]
                                self.ax.plot(x,y)
                                self.axes()
                                self.canvas.draw()
                            else: continue
                        
            if self.plotType ==1:
                for dF in range(len(self.dataFiles)):
                                       
                    x = np.arange(0,self.controlMeans.size*max(self.interval),max(self.interval))
                    y = self.controlMeans
                    
                    self.clear()
                    plt.errorbar(x,y,color = 'b', yerr = self.controlSEM)
                    
                    self.axes()
                    self.canvas.draw()
        logging.debug('Plot() Completed')
                    
    def peaks(self):
       
        peakList=[]
        halfList = []
        for dF in range(len(self.dataFiles)):
            for s in range(self.dmgRois[dF]):
                if self.Chk[dF][s].isChecked()==True:
                    if (roiInfo[dF][s,2]!=0.0):
                        peakList.append(roiInfo[dF][s,2])
                    if (roiInfo[dF][s,3]!=0.0):
                        halfList.append(roiInfo[dF][s,3])
        meanPeak = np.mean(peakList)
        peakSEM = stats.sem(peakList)
        meanHalf = np.mean(halfList)
        halfSEM = stats.sem(halfList)
       
        self.peakTime = float(format(meanPeak, '.2f'))
        self.halfTime = float(format(meanHalf, '.2f'))
        self.peakError = float(format(peakSEM, '.2f'))
        self.halfError = float(format(halfSEM, '.2f'))
        self.textbox.setText('Mean Time to Peak Recruitment:  ' + str((self.peakTime)) + "      SEM: " + str(self.peakError))
        self.textbox2.setText('Mean Half-Life of Recruitment: ' + str((self.halfTime))+ "       SEM: "  + str(self.halfError))
        logging.debug('Peaks() Completed')
    
    def clear(self):
        self.ax.cla()
        self.canvas.draw()
        logging.debug('clear() Completed')
    
    def axes(self):
        
        durations =  []
        for dF in range(len(self.dataFiles)):
            durations.append(math.ceil(data[dF][:,self.dataStructure['Time'],:].max()/1000))
        
        maxXRange = max(durations)
        maxYRanges = []
        if self.scheme ==1:
            for dF in range(len(self.dataFiles)):
                maxYRanges.append(math.ceil(data[dF][:,self.dataStructure['rawStim'],:].max()/100)*100)    
            
            
            
        if self.scheme == 2:
            for dF in range(len(self.dataFiles)):
                maxYRanges.append( math.ceil(data[dF][:,self.dataStructure['refStim'],:].max()/100)*100)    
            
            
        if self.scheme == 3:
            for dF in range(len(self.dataFiles)):
                maxYRanges.append(math.ceil(data[dF][:,self.dataStructure['refDiv'],:].max()))    
           
        
        if self.scheme == 4:
            for dF in range(len(self.dataFiles)):
                maxYRanges.append(math.ceil(data[dF][:,self.dataStructure['refSub'],:].max()))
                           
            maxYRanges.append(1.2)
        if self.scheme == 5:
            maxYRanges.append(1.2)
                
        if self.scheme ==6:
            for dF in range(len(self.dataFiles)):
                maxYRanges.append(math.ceil(data[dF][:,self.dataStructure['normFirst'],:].max()))
        maxYRange =  max(maxYRanges)  
        self.ax.axis((-20,maxXRange,0,maxYRange)) 
        logging.debug('Axes() Completed')

    def selectionchange(self):
        self.controlType = self.cb.currentIndex()
        self.clear()
        self.plot()
        logging.debug('selectionchange() Completed')
        
    def selectionchange2(self):
        self.plotType = self.cb2.currentIndex()
        self.clear()
        self.plot()
        logging.debug('selectionchange2() Completed')
        
    def btnstate(self):
        self.clear()
        self.plot()  
        logging.debug('btnstate() Completed')
        
    def radio_button_clicked(self):
        self.scheme = self.buttongroup.checkedId()
        self.peaks()
        self.clear()
        self.plot()                        
        logging.debug('radio_button_clicked() Completed')
          
if __name__ == '__main__':
    app = 0    
    app = QtWidgets.QApplication(sys.argv)
    
    main = Window(None)
    logging.debug('window() called')
    main.show()
    logging.debug('main.show() called')
    sys.exit(app.exec_())
    
   # -*- coding: utf-8 -*-


